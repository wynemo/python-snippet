#coding:utf-8
from pathlib import Path
import os
import os.path
import sys

import xlwt
from openpyxl import load_workbook

def filter_path(path):
    path = str(path)
    if path.find('2018') != -1:
        return True
    if path.find('2019') != -1:
        return True
    if path.find('2020') != -1:
        return True
    if path.find('2021') != -1:
        return True
    return False

class Excel():
    def __init__(self, folder_name, full_path):
        if folder_name.startswith('\\'):
            folder_name = folder_name[1:]
        self.folder_name = folder_name
        self.full_path = full_path


def main():
    if len(sys.argv) < 2:
        print('please specify a folder')
        os._exit(0)
    folder = sys.argv[1]
    print('folder is', folder)
    all_path = []
    for path in Path(folder).rglob('*附表1*.xls*'):
        full_path = os.path.join(path.parent, path.name)
        folder_name = str(path.parent).replace(folder, '', 1)
        if filter_path(path.name):
            continue
        else:
            excel = Excel(folder_name, full_path)
            all_path.append(excel)
    for each in all_path:
        handle_excel(each)
        break
        # print(path.name, str(path.parent).replace(folder, '', 1))

def handle_excel(excel: Excel):
    wb2 = load_workbook(excel.full_path)
    if len(wb2.sheetnames) and wb2.sheetnames[0].startswith('Sheet'):
        handle_sheet(excel, wb2[wb2.sheetnames[0]])

def handle_sheet(excel: Excel, sheet):
    # print(excel.full_path, sheet)
    values = {'base': [], 'bonus': [], 'other': []}
    stage = None
    for row in sheet.iter_rows(values_only=True):
        if filter_useless_row(row):
            continue
        if stage is None:
            if find_base(row[0]):
                stage = 'base'
            else:
                continue
        elif stage == 'base':
            if not filter_useless_row(row):
                values['base'].append(row)
        elif stage == 'bonus':
            if not filter_useless_row(row):
                values['bonus'].append(row)
        elif stage == 'other':
            if not filter_useless_row(row):
                values['bonus'].append(row)
    print(values) 
        

def find_base(value: str):
    return '基本工资' in value

def find_bonus(value: str):
    return '津贴补贴' in value

def find_other(value: str):
    return '其他' in value


def filter_empty_value_row(row):
    if not row[7]: # 发放人数
        return True
    return False


def filter_useless_row(row):
    if not row:
        return True
    if not row[0]:
        return True
    if not isinstance(row[0], str):
        return True
    header = row[0].strip()
    if header in ('总计', '合计', '小计'):
        return True
    return False

if __name__ == '__main__':
    main()

